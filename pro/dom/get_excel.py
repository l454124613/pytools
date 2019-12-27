# -*- coding:utf-8 -*-
# Author:lixuecheng
from openpyxl import load_workbook
from core.logger import logger, log
from core.do_mysql import do_mysql
from config import sys_config
from util.check import check_none_kong


@log
def read_excel(file_path):
    logger.info('---------------------读取excel---------------------')
    wb = load_workbook(file_path)
    # 将来的全局变量
    g = {}
    excel_inf = {'config': []}

    # 读取配置信息
    if sys_config.config_table in wb.sheetnames:
        sheet = wb[sys_config.config_table]
        g['当前运行环境'] = sheet.cell(row=1, column=2).value
        # g['运行要求'] = sheet.cell(row=1, column=3).value
        # if g['运行要求'] is None or g['运行要求'] == '':
        #     raise Exception('请设置是否流程还是单独运行')
        for i in range(3, sheet.max_row + 1):
            env = sheet.cell(row=i, column=1).value
            item = sheet.cell(row=i, column=2).value
            para = sheet.cell(row=i, column=3).value
            va = sheet.cell(row=i, column=4).value
            com = sheet.cell(row=i, column=5).value
            if com is None:
                com = ''
            num = check_none_kong(env, item, para, va)
            if num == 4:
                continue
            elif num == 0:
                if g.get(env) is None:
                    g[env] = {}
                if item == '字符串':
                    g[env][para] = str(va).strip()
                    logger.info(g[env][para] + ',配置值读取成功')
                elif item == '数组':
                    login_info = va.replace('；', ';').split(';')
                    if len(login_info) == 2:
                        g[env][para] = (login_info[0], login_info[1])
                        logger.info(str(va) + ',配置数组读取成功')
                    else:
                        raise Exception('登录名参数值不正确，请使用：用户名;密码，当前是：' + str(va))
                elif item == 'mysql':
                    database = va.replace('；', ';').split(';')
                    if len(database) == 5:
                        g[env][para] = do_mysql(*database)
                        logger.info(str(va) + ',mysql连接成功')
                    else:
                        raise Exception('数据库参数值不正确，请使用：路径；端口；用户名；密码；数据库名称，当前是：' + str(va))
                excel_inf['config'].append({'env': env, 'type': item, 'name': para, 'value': va, 'com': com})
                # elif item == 'post':
                #     post_info = va.replace('；', ';').split(';')
                #     if len(post_info) == 3:
                #         g[env][para] = do_request(para, post_info, is_func=False)
                #         logger.info(str(va) + ',post读取成功')
                # elif item == 'get':
                #     g[env][para] = do_request(para, va, is_func=False)
                #     logger.info(str(va) + ',get读取成功')


            else:
                logger.warn("运行环境:%s,设置项目:%s,参数名称:%s,参数值:%s,有值为None" % (env, item, para, va))
    else:
        raise Exception('无配置文件，无法执行,配置sheet名为"' + sys_config.config_table + '"')

    # d3 = {}
    # d4 = 0
    # # ds = []
    # r4s = {'sheets': {}}
    # nn = 0

    # 所有用例的信息
    cases_dict = {}
    sheet_cases_map = {}
    # 添加配置文件到忽略列表
    sys_config.ignore_table.append(sys_config.config_table)
    for i in wb.sheetnames:
        if i in sys_config.ignore_table:
            continue
        else:
            #
            sheet = wb[i]
            i = i.strip()
            sheet_cases_map[i] = []
            for j in range(2, sheet.max_row + 1):
                name = sheet.cell(row=j, column=int(sys_config.name)).value
                method = sheet.cell(row=j, column=int(sys_config.method)).value
                ip = sheet.cell(row=j, column=int(sys_config.host)).value
                path = sheet.cell(row=j, column=int(sys_config.path)).value
                da = sheet.cell(row=j, column=int(sys_config.fixed_body)).value
                header = sheet.cell(row=j, column=int(sys_config.header_content_type)).value
                pa = sheet.cell(row=j, column=int(sys_config.parameter)).value
                forward = sheet.cell(row=j, column=int(sys_config.pre_cases)).value
                preaction = sheet.cell(row=j, column=int(sys_config.pre_operation)).value
                reaction = sheet.cell(row=j, column=int(sys_config.post_operation)).value
                cvalue = sheet.cell(row=j, column=int(sys_config.response_check)).value
                csql = sheet.cell(row=j, column=int(sys_config.sql_check)).value
                isrun = sheet.cell(row=j, column=int(sys_config.is_run)).value
                num = check_none_kong(name, method, ip, path, da, header, pa, forward, preaction, reaction, cvalue,
                                      csql)
                num1 = check_none_kong(name, method, ip, path, cvalue, isrun)
                # 排除空行
                if num == 12 or isrun is None or isrun == 'stop':
                    continue
                # 必填项填写且可执行的
                elif num1 == 0 and isrun != 'stop':
                    # 校验重复名称
                    if name in cases_dict:
                        raise Exception('用例名称重复：' + name)
                    else:
                        sheet_cases_map[i].append(name)
                        cases_dict[name] = {'sheetname': str(i), 'casename': name, 'method': method, 'ip': ip,
                                            'path': path, 'da': da, 'header': header, 'pa': pa, 'forward': forward,
                                            'pre': preaction, 're': reaction, 'cvalue': cvalue, 'csql': csql,
                                            'row_num': j, 'status': isrun}

                else:
                    logger.warn(i + '表的' + str(j) + '行，必填字段未填，请检查')
                    continue

    logger.info('###################读取excel,共有{}条用例####################'.format(str(len(cases_dict))))
    wb.close()
    return g, sheet_cases_map, cases_dict, excel_inf
    # d3.update(d2)

    # ws = wb[i]

    # print(ws)

# wb.close()
# print('获取用例数量：' + str(len(d3)))
# if d4 > 0:
#     print('跳过用例数量：' + str(d4))
# logger.info('###################读取excel####################')
# # g['sheets'] = ds
# g['row4sheet'] = r4s
# # g['cases_num'] = str(len(d3))
# return g, d3

# a = read_excel(r"E:\SVN_project\others\接口自动化\v2\test43.xlsx")
# print(a)
