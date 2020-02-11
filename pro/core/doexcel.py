# -*- coding:utf-8 -*-
# Author:lixuecheng

import openpyxl
import xlwt
import xlrd
from pro.core.logger import log, logger
import os

from xlutils.copy import copy


class DoExcel:
    '''
    此方法用于excel常用的读数据和写数据的操作

    @params path excel文件的路径，只支持xlsx和xls的格式，其他格式无法使用

    @params need_exception 默认true，执行时，raise 异常；false，执行时，不中断，异常在方法后返回值（bool，str）
    '''

    def __init__(self, path, need_exception=True):
        '''
        @params path excel文件的路径，只支持xlsx和xls的格式，其他格式无法使用

        @params need_exception 默认true，执行时，raise 异常；false，执行时，不中断，异常在方法后返回值（bool，str）
        '''
        self.need_exception = need_exception
        self.path = path
        self.file_dir = os.path.dirname(path)
        self.file_name = os.path.basename(path)
        self.file_type = os.path.splitext(self.file_name)[1]
        if self.file_type == '' or self.file_type not in ['.xlsx', '.xls']:
            self.file_status = False
            raise Exception('当前只支持xlsx和xls,请检查格式')
        else:
            self.file_status = True
        self.file_is_exists = os.path.exists(path)
        self.wb = None
        self.file_rd_status = True
        self.file_wt_status = True
        self.is_wt = False
        self.sheet_name = {}

    def read(self, sheet, row, column):
        '''
        @params sheet 表格的名字

        @params row 行号，从1开始

        @params column 列号，从1开始

        @return (bool,str) true 返回读到的值  false 返回异常情况
        '''
        if self.is_wt:
            if self.need_exception:

                raise Exception('文件已被修改，无法读取,不能read和write同时使用')
            else:

                return False, '文件已被修改，无法读取,不能read和write同时使用'
        if self.file_status and self.file_rd_status and self.file_is_exists:
            try:
                row = int(row)
                column = int(column)
            except:
                if self.need_exception:

                    raise Exception('输入行或列需要数字')
                else:

                    return False, '输入行或列需要数字'
            if self.file_type == '.xlsx':
                try:
                    if self.wb is None:
                        self.wb = openpyxl.load_workbook(self.path)
                    if sheet in self.wb.sheetnames:
                        sh = self.wb[sheet]

                        if row is None or column is None:
                            return False, "请输入行和列"
                        elif row > sh.max_row or column > sh.max_column:
                            return False, "输入行或列超过最大值"

                        else:
                            v = sh.cell(row, column).value
                            return True, v or ''

                    else:
                        if self.need_exception:

                            raise Exception('没有指定的sheet:'+sheet)
                        else:

                            return False, '没有指定的sheet:'+sheet

                except Exception as e:
                    if self.need_exception:
                        self.file_rd_status = False
                        raise e
                    else:
                        self.file_rd_status = False
                        return False, str(e)

            else:
                try:
                    if self.wb is None:
                        self.wb = xlrd.open_workbook(self.path)
                        inn = 0
                        for i in self.wb.sheet_names():

                            self.sheet_name[i] = inn
                            inn += 1

                    if sheet in self.wb.sheet_names():
                        sh = self.wb.sheet_by_name(sheet)
                        # print(dir(sh))

                        if row is None or column is None:
                            return False, "请输入行和列"
                        elif row > sh.nrows or column > sh.ncols:
                            return False, "输入行或列超过最大值"
                        elif row < 1 or column < 1:
                            return False, "输入行或列必须大于0的整数"

                        else:
                            # print(sh.nrows)
                            # print(sh.utter_max_cols)
                            v = sh.cell(row-1, column-1).value
                            return True, v

                    else:
                        if self.need_exception:

                            raise Exception('没有指定的sheet:'+sheet)
                        else:

                            return False, '没有指定的sheet:'+sheet

                except Exception as e:
                    if self.need_exception:
                        self.file_rd_status = False
                        raise e
                    else:
                        self.file_rd_status = False
                        return False, str(e)

        else:
            if self.need_exception:
                self.file_rd_status = False
                raise Exception('文件状态不正确或者不存在：'+self.path)
            else:
                self.file_rd_status = False
                return False, '文件状态不正确或者不存在：'+self.path

    def close(self):
        '''
        在结束时，如果有写操作，必须关闭，使用此方法
        '''
        try:
            if self.is_wt:

                self.wb.save(self.path)
                self.is_wt = False
            self.wb.close()
        except:
            pass

    def __del__(self):

        if self.is_wt:
            print('请确认excel没有打开中并使用close()方式关闭，不然文件会损坏并报错')

    def write(self, sheet, row, column, value):
        '''
        @params sheet 表格的名字

        @params row 行号，从1开始

        @params column 列号，从1开始

        @params value 值

        @return (bool,str) true 无值  false 返回异常情况
        '''
        if self.file_status and self.file_wt_status:
            try:
                row = int(row)
                column = int(column)
            except:
                if self.need_exception:

                    raise Exception('输入行或列需要数字')
                else:

                    return False, '输入行或列需要数字'
            if self.file_type == '.xlsx':
                if self.file_is_exists:
                    if self.wb is None:
                        self.wb = openpyxl.load_workbook(self.path)
                    if sheet in self.wb.sheetnames:
                        sh = self.wb[sheet]

                    else:
                        sh = self.wb.create_sheet(title=sheet)

                    sh.cell(row, column, value=str(value))
                    self.is_wt = True
                    return True, ''
                else:
                    self.wb = openpyxl.Workbook()
                    self.wb.remove_sheet(self.wb['Sheet'])
                    sh = self.wb.create_sheet(title=sheet)
                    sh.cell(row, column, value=str(value))
                    self.is_wt = True
                    # self.wb.save(self.path)
                    return True, ''
            else:
                if self.file_is_exists:
                    if self.wb is None:
                        self.wb = xlrd.open_workbook(self.path)
                        inn = 0
                        for i in self.wb.sheet_names():

                            self.sheet_name[i] = inn
                            inn += 1

                        self.wb = copy(self.wb)
                    # print(dir(self.wb))
                    # print(self.sheet_name)
                    if sheet in self.sheet_name:

                        sh = self.wb.get_sheet(self.sheet_name[sheet])

                    else:
                        sh = self.wb.add_sheet(sheet)
                        self.sheet_name[sheet] = len(self.sheet_name)

                    sh.write(row, column, str(value))
                    self.is_wt = True
                    return True, ''
                else:
                    self.wb = xlwt.Workbook()
                    self.file_is_exists = True
                    # self.wb.remove_sheet(self.wb['Sheet'])
                    sh = self.wb.add_sheet(sheet)
                    self.sheet_name[sheet] = len(self.sheet_name)
                    sh.write(row, column, str(value))
                    self.is_wt = True
                    # self.wb.save(self.path)
                    return True, ''

        else:
            if self.need_exception:
                self.file_wt_status = False
                raise Exception('文件状态不正确：'+self.path)
            else:
                self.file_wt_status = False
                return False, '文件状态不正确：'+self.path

    def writes(self, sheet, values):
        '''
        @params sheet 表格的名字

        @params values 值 格式[(x,y,v)]

        @return (bool,str) true 无值  false 返回异常情况

        '''
        if self.file_status and self.file_wt_status:
            v = []
            try:
                for value in values:
                    row = int(value[0])
                    column = int(value[1])
                    v.append({'row': row, 'column': column, 'value': value[2]})
            except:
                if self.need_exception:

                    raise Exception('输入行或列需要数字和值')
                else:

                    return False, '输入行或列需要数字和值'
            if self.file_type == '.xlsx':
                if self.file_is_exists:
                    if self.wb is None:
                        self.wb = openpyxl.load_workbook(self.path)
                    if sheet in self.wb.sheetnames:
                        sh = self.wb[sheet]

                    else:
                        sh = self.wb.create_sheet(title=sheet)

                    for i in v:
                        sh.cell(i['row'], i['column'], value=str(i['value']))
                    self.is_wt = True
                    return True, ''
                else:
                    self.wb = openpyxl.Workbook()
                    self.wb.remove_sheet(self.wb['Sheet'])
                    sh = self.wb.create_sheet(title=sheet)
                    for i in v:
                        sh.cell(i['row'], i['column'], value=str(i['value']))
                    self.is_wt = True
                    # self.wb.save(self.path)
                    return True, ''
            else:
                if self.file_is_exists:
                    if self.wb is None:
                        self.wb = xlrd.open_workbook(self.path)
                        inn = 0
                        for i in self.wb.sheet_names():

                            self.sheet_name[i] = inn
                            inn += 1

                        self.wb = copy(self.wb)
                    # print(dir(self.wb))
                    # print(self.sheet_name)
                    if sheet in self.sheet_name:

                        sh = self.wb.get_sheet(self.sheet_name[sheet])

                    else:
                        sh = self.wb.add_sheet(sheet)
                        self.sheet_name[sheet] = len(self.sheet_name)

                    for i in v:
                        sh.write(i['row']-1, i['column']-1, str(i['value']))

                    self.is_wt = True
                    return True, ''
                else:
                    self.wb = xlwt.Workbook()
                    self.file_is_exists = True
                    # self.wb.remove_sheet(self.wb['Sheet'])
                    sh = self.wb.add_sheet(sheet)
                    self.sheet_name[sheet] = len(self.sheet_name)
                    for i in v:
                        sh.write(i['row']-1, i['column']-1, str(i['value']))
                    self.is_wt = True
                    # self.wb.save(self.path)
                    return True, ''

        else:
            if self.need_exception:
                self.file_wt_status = False
                raise Exception('文件状态不正确：'+self.path)
            else:
                self.file_wt_status = False
                return False, '文件状态不正确：'+self.path


# d = DoExcel(r'C:\Users\lixuecheng\Desktop\时间管理测试范围 - 副本.xlsx')
# # c = d.read('ucb', 1, 1)
# d.write('ucb',30,7,12322)
# d.write('ucb',30,8,112322)
# d.write('ucb1',30,8,112322)
# d.close()
# c = d.read('ucb', 30, 7)
# d.write('ttt',2,2,13)
# c=d.read('ucb',3,3)
# print(c)
# d=DoExcel(r'C:\Users\lixuecheng\Desktop\aa111.xls')
# d.writes('tt',[(1,1,22),(2,2,'hahacz这些年')])
# d.close()
